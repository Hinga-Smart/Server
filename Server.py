from flask import Flask, request, jsonify, send_file, render_template
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import traceback

app = Flask(__name__)
file_name = "moisture_sensor_data.xlsx"
log_file = "error_log.txt"

DRY_THRESHOLD = 300
WET_THRESHOLD = 700

# --- Utility setup ---

def log_error(message):
    with open(log_file, "a") as f:
        f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {message}\n")

def getState(moisture):
    if moisture < DRY_THRESHOLD:
        return "DRY"
    elif moisture > WET_THRESHOLD:
        return "WET"
    else:
        return "MODERATE"

# Create Excel file if missing
if not os.path.isfile(file_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Moisture soil data"
    ws.append(["Timestamp", "Moisture", "State"])
    wb.save(file_name)

# --- Routes ---

@app.route('/')
def home():
    """Render dashboard HTML page"""
    # Load data for immediate chart rendering
    wb = load_workbook(file_name)
    ws = wb.active
    rows = list(ws.values)
    headers = rows[0]
    data = [dict(zip(headers, r)) for r in rows[1:]]
    return render_template('index.html', data=data)

@app.route('/data', methods=['POST'])
def sensorData():
    """Receive data from ESP sensor"""
    try:
        results = request.get_json(force=True, silent=True)
        log_error(f"Raw data: {request.data}")
        log_error(f"Parsed JSON: {results}")

        if not results or 'moisture' not in results:
            return {"status": "Invalid JSON"}, 400

        moisture = int(results['moisture'])
        state = getState(moisture)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        wb = load_workbook(file_name)
        ws = wb.active
        ws.append([timestamp, moisture, state])
        wb.save(file_name)

        print(f"Data received: {timestamp}, Moisture: {moisture}, State: {state}")
        log_error(f"Data recorded successfully: {timestamp}, {moisture}, {state}")

        return {"status": "Data recorded successfully"}

    except Exception as e:
        error_message = f"Error: {e}\n{traceback.format_exc()}"
        log_error(error_message)
        print(error_message)
        return {"status": "Server error"}, 500

@app.route('/latest', methods=['GET'])
def latestData():
    """Return the most recent reading as JSON"""
    if not os.path.isfile(file_name):
        return jsonify({}), 404
    wb = load_workbook(file_name)
    ws = wb.active
    rows = list(ws.values)
    headers = rows[0]
    latest = dict(zip(headers, rows[-1]))
    return jsonify(latest)

@app.route('/download', methods=['GET'])
def downloadFile():
    """Download the full Excel file"""
    if not os.path.isfile(file_name):
        return {"status": "No data file found"}, 404
    return send_file(file_name, as_attachment=True)

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 1000))
    app.run(host="0.0.0.0", port=port)